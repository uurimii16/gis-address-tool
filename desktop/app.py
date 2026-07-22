# -*- coding: utf-8 -*-
"""
GIS 주소 변환기 (데스크톱) · v8  ── 대용량 안정 병렬 + 원본열 보존 + 파일 미리보기
 기능 ① 주소 → PNU   (정제주소·본번·부번, 공시지가 옵션)
 기능 ② 주소 → 좌표  (위경도 또는 중부원점TM)
 기능 ③ QGIS 레이어  (포인트·필지 경계 GeoJSON)

대용량 안정성(핵심):
 · 동시 조회 수는 config.txt 의 WORKERS(기본 4). 너무 크면 VWorld가 연결을 끊어 실패↑ → 작게 시작.
 · 회로차단기 없음. 연결실패/타임아웃/5xx는 지수 백오프로 최대 5회 재시도(일시적 조임을 넘긴다).
 · 전 요청 공통 속도 상한(과도한 동시 접속 방지).
읽기/출력:
 · 미리보기·열선택은 앞 HEAD_ROWS 행만(큰 파일도 즉시). 변환은 read_only 스트리밍.
 · 출력 = 원본 열 전체 + 결과 열(입력주소·PNU·정제주소·본번·부번·상태[+공시지가]). PNU는 텍스트라 지수표기(E+) 없음.
데스크톱은 로컬 실행이라 연결 끊김·타임아웃이 없어 대용량(수만 행) 작업에 적합. 설정은 같은 폴더 config.txt.
"""
import os
import io
import re
import sys
import csv
import json
import time
import random
import threading
import webbrowser
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import requests
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl import load_workbook, Workbook

GEOCODE_URL = "https://api.vworld.kr/req/address"
DATA_URL = "https://api.vworld.kr/req/data"

WORKERS_DEFAULT = 4    # 동시에 물어보는 주소 개수(기본). config.txt WORKERS 로 조절.
POOL_SIZE = 16         # 커넥션 풀 상한(WORKERS 를 나중에 키워도 되도록 넉넉히)
MAX_ROWS = 500_000     # 안전을 위한 최대 처리 행 수
MAX_COLS = 60          # 읽어들일 최대 열 수
HEAD_ROWS = 300        # 미리보기·열 자동인식에 쓸 앞부분 행 수
REQ_TIMEOUT = 15       # 한 번 호출을 기다리는 최대 시간(초)
GEO_RETRIES = 5        # 일시적 실패 시 재시도 횟수
REQ_MIN_INTERVAL = 0.05  # 요청 시작 간 최소 간격(초) — 과도한 동시 접속 방지(대략 초당 ~20건 상한)

SIDO = ["서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시", "대전광역시",
        "울산광역시", "세종특별자치시", "경기도", "강원특별자치도", "강원도", "충청북도",
        "충청남도", "전북특별자치도", "전라북도", "전라남도", "경상북도", "경상남도",
        "제주특별자치도", "충북", "충남", "전북", "전남", "경북", "경남", "경기", "강원", "제주"]

CSV_ENCODINGS = [
    ("자동 감지", None),
    ("CP949 / EUC-KR (공공데이터·엑셀 저장 CSV)", "cp949"),
    ("UTF-8", "utf-8-sig"),
]
AUTO_TRY = ["utf-8-sig", "cp949", "euc-kr", "latin1"]

# ----- 밝은 테마 색상 -----
BG = "#e1e9ef"
CARD = "#ffffff"
INK = "#2c3e50"
MUTED = "#7e8ea0"
ACCENT = "#5b8db5"
ACCENT_D = "#487aa3"
SOFT = "#cce2f3"
WARM = "#efe2df"
TAB_OFF_BG = "#f3f6f9"
LOGBG = "#f5f8fb"
LINE = "#d4dde6"

# UI 폰트: 세련된 한글 폰트를 우선 사용하고, 없으면 자동으로 대체(맑은 고딕).
UI_FONT = "맑은 고딕"   # _pick_font() 로 실행 시 교체
FONT_PREFER = [
    "KoPubWorld돋움체 Medium", "KoPubWorld돋움체 Light", "KoPubWorld돋움체",
    "KoPub돋움체_Pro Medium", "KoPub돋움체_Pro Light", "KoPub돋움체",
    "Pretendard Variable", "Pretendard", "본고딕", "Noto Sans KR",
    "Malgun Gothic", "맑은 고딕",
]

PICK_TEXT = "📂  파일 선택 (.xlsx · .csv)  ·  또는 파일을 여기로 끌어다 놓기"

DEFAULT_CONFIG = """# ===== 변환기 설정 =====
# VWorld 인증키 (필수) — 각자 vworld.kr 에서 무료로 발급받아 붙여넣으세요.
API_KEY=여기에_VWorld_인증키_붙여넣기

# 공시지가·필지 조회 시 데이터 API 도메인 (키 등록 URL과 동일하게)
DOMAIN=localhost

# 동시에 조회하는 개수 (많을수록 빠르지만 너무 크면 VWorld가 연결을 끊어 실패가 늘어요).
# '통신실패'가 많으면 4 -> 3 으로 줄이고, 계속 안정적이면 6~8 로 올려 보세요.
WORKERS=4
"""

CRS_OPTIONS = {
    "위경도 (EPSG:4326)": ("EPSG:4326", "경도(X)", "위도(Y)"),
    "중부원점TM (EPSG:5186)": ("EPSG:5186", "X(TM)", "Y(TM)"),
}

# 커넥션 재사용용 세션 (TLS 핸드셰이크 재활용). 기본 UA 를 막는 방화벽 대비 UA 지정.
_session = requests.Session()
_session.headers.update({"User-Agent": "gis-address-tool/1.0 (+github.com/uurimii16/gis-address-tool)"})
_adapter = requests.adapters.HTTPAdapter(
    pool_connections=POOL_SIZE, pool_maxsize=POOL_SIZE, max_retries=0)
_session.mount("https://", _adapter)
_session.mount("http://", _adapter)

# 전역 요청 속도 상한: 여러 스레드가 동시에 몰려도 요청 '시작' 간격을 살짝 벌려 준다.
_rate_lock = threading.Lock()
_last_req = {"t": 0.0}


def _rate_gate():
    with _rate_lock:
        now = time.monotonic()
        wait = REQ_MIN_INTERVAL - (now - _last_req["t"])
        if wait > 0:
            time.sleep(wait)
            now = time.monotonic()
        _last_req["t"] = now


def _pick_font(root):
    """세련된 한글 폰트를 우선 선택하고, 설치돼 있지 않으면 맑은 고딕으로 대체."""
    try:
        import tkinter.font as tkfont
        avail = set(tkfont.families(root))
    except Exception:
        return "맑은 고딕"
    for f in FONT_PREFER:
        if f in avail:
            return f
    return "맑은 고딕"


# ---------- 유틸 ----------
def app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(name):
    base = getattr(sys, "_MEIPASS", app_dir())
    return os.path.join(base, name)


def config_path():
    return os.path.join(app_dir(), "config.txt")


def load_config():
    path = config_path()
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            f.write(DEFAULT_CONFIG)
    cfg = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            cfg[k.strip()] = v.strip()
    return cfg


def set_config_value(key, value):
    """config.txt 의 한 항목(예: API_KEY)만 갱신하고 나머지(주석·다른 설정)는 보존한다."""
    path = config_path()
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            f.write(DEFAULT_CONFIG)
    with open(path, "r", encoding="utf-8") as f:
        lines = f.read().splitlines()
    found = False
    for i, line in enumerate(lines):
        s = line.strip()
        if s and not s.startswith("#") and "=" in s and s.split("=", 1)[0].strip() == key:
            lines[i] = f"{key}={value}"; found = True; break
    if not found:
        lines.append(f"{key}={value}")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def idx_to_col(i):
    s = ""
    while i > 0:
        i, rem = divmod(i - 1, 26)
        s = chr(65 + rem) + s
    return s


def parse_bonbu(pnu):
    if pnu and len(pnu) == 19:
        return int(pnu[11:15]), int(pnu[15:19])
    return None, None


# ---------- VWorld ----------
def _vworld_json(url, params, timeout=REQ_TIMEOUT):
    """(json, None) 또는 (None, 사람이 읽을 실패사유)."""
    try:
        r = _session.get(url, params=params, timeout=timeout)
    except requests.RequestException as e:
        return None, "연결실패:" + type(e).__name__
    if r.status_code != 200:
        return None, f"HTTP{r.status_code}"
    try:
        return r.json(), None
    except ValueError:
        body = (r.text or "").strip()
        return None, ("빈응답(IP제한 의심)" if not body
                      else "비정상응답:" + body[:60].replace("\n", " "))


def validate_key(api_key):
    """인증키 유효성 확인 → 'ok' · 'invalid' · 'network'."""
    j, err = _vworld_json(GEOCODE_URL, {
        "service": "address", "request": "getcoord", "version": "2.0",
        "crs": "EPSG:4326", "address": "서울특별시 중구 세종대로 110",
        "type": "ROAD", "format": "json", "key": api_key}, timeout=8)
    if err is not None:
        return "network", err
    resp = j.get("response", {}) if isinstance(j, dict) else {}
    if resp.get("status") == "OK":
        return "ok", ""
    err_obj = resp.get("error") or {}
    code = str(err_obj.get("code", "")).upper()
    text = str(err_obj.get("text", ""))
    if ("KEY" in code or code in {"020", "021"} or "인증키" in text or "등록되지" in text):
        return "invalid", f"{code} {text}".strip()
    return "ok", ""


def _backoff(attempt):
    """지수 백오프 + 지터(초). 일시적 조임을 넘기기 위한 대기."""
    return min(8.0, 0.5 * (2 ** attempt)) + random.uniform(0, 0.3)


def geocode(addr, api_key, crs="EPSG:4326", retries=GEO_RETRIES):
    last = "실패"
    for attempt in range(retries):
        _rate_gate()
        j, err = _vworld_json(GEOCODE_URL, {
            "service": "address", "request": "getcoord", "version": "2.0",
            "crs": crs, "address": addr, "type": "PARCEL",
            "format": "json", "key": api_key})
        if err is None:
            resp = j.get("response", {})
            if resp.get("status") == "OK":
                try:
                    s = resp["refined"]["structure"]
                    pt = resp["result"]["point"]
                    return (s.get("level4LC", ""), pt["x"], pt["y"],
                            resp["refined"].get("text", ""), "OK")
                except (KeyError, TypeError):
                    return None, None, None, None, "주소인식실패"
            code = (resp.get("error") or {}).get("code", "")
            if str(code).upper() == "INVALID_KEY":
                return None, None, None, None, "인증키오류"
            return None, None, None, None, "주소인식실패"
        # 연결실패·타임아웃·5xx·빈응답 등 일시적 → 잠깐 쉬고 재시도
        last = err
        if attempt < retries - 1:
            time.sleep(_backoff(attempt))
    return None, None, None, None, f"통신실패:{last}"


def get_parcel(x, y, api_key, domain="localhost", retries=4):
    for attempt in range(retries):
        _rate_gate()
        j, err = _vworld_json(DATA_URL, {
            "service": "data", "request": "GetFeature", "data": "LP_PA_CBND_BUBUN",
            "version": "2.0", "geomFilter": f"POINT({x} {y})", "crs": "EPSG:4326",
            "format": "json", "size": "1", "domain": domain, "key": api_key})
        if err is None:
            resp = j.get("response", {})
            if resp.get("status") != "OK":
                return None, None
            try:
                feat = resp["result"]["featureCollection"]["features"][0]
                return feat.get("geometry"), feat.get("properties", {})
            except (KeyError, IndexError, TypeError):
                return None, None
        if attempt < retries - 1:
            time.sleep(_backoff(attempt))
    return None, None


def work(addr, api_key, crs, need_parcel, domain):
    """병렬 실행 단위: 한 주소의 지오코딩(+필요시 필지)을 묶어 처리."""
    pnu, x, y, refined, status = geocode(addr, api_key, crs)
    item = {"addr": addr, "status": status, "pnu": pnu, "x": x, "y": y, "refined": refined}
    if need_parcel and status == "OK" and x:
        item["geom"], item["props"] = get_parcel(x, y, api_key, domain)
    return item


# ---------- 파일 읽기 ----------
def decode_csv(raw, enc):
    tries = [enc] if enc else AUTO_TRY
    last = None
    for e in tries:
        try:
            return raw.decode(e), e
        except Exception as ex:
            last = ex
    raise last or ValueError("인코딩을 인식하지 못했습니다.")


def _row_cells(row):
    return ["" if v is None else str(v).strip() for v in row[:MAX_COLS]]


def xlsx_sheet_names(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def read_head(path, ext, sheet, enc, raw=None, n=HEAD_ROWS):
    """미리보기·열 선택용으로 앞 n행만 grid 로 읽는다 → (grid, 사용인코딩)."""
    used = enc
    grid = []
    if ext == "csv":
        text, used = decode_csv(raw, enc)
        for row in csv.reader(io.StringIO(text)):
            grid.append(_row_cells(row))
            if len(grid) >= n:
                break
    else:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
            for row in ws.iter_rows(values_only=True):
                grid.append(_row_cells(row))
                if len(grid) >= n:
                    break
        finally:
            wb.close()
    return grid, used


def _iter_full(path, ext, sheet, enc, raw=None):
    if ext == "csv":
        text, _ = decode_csv(raw, enc)
        for row in csv.reader(io.StringIO(text)):
            yield _row_cells(row)
    else:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
            for row in ws.iter_rows(values_only=True):
                yield _row_cells(row)
        finally:
            wb.close()


def read_full_grid(path, ext, sheet, enc, raw=None):
    """전체 행을 grid(문자열 2차원)로. 원본 열을 결과에 함께 담기 위해 필요."""
    grid, truncated = [], False
    for row in _iter_full(path, ext, sheet, enc, raw):
        grid.append(row)
        if len(grid) >= MAX_ROWS:
            truncated = True
            break
    return grid, truncated


def collect_addresses(path, ext, sheet, enc, sel, prefix, start_row, raw=None):
    """레이어(③)용: 원본 열 없이 주소 문자열만 스트리밍으로 모은다."""
    addrs, truncated = [], False
    for i, row in enumerate(_iter_full(path, ext, sheet, enc, raw), start=1):
        if i < start_row:
            continue
        addrs.append(full_addr_row(row, sel, prefix))
        if len(addrs) >= MAX_ROWS:
            truncated = True
            break
    return addrs, truncated


def save_with_original(grid_full, base_width, start_row, result_by_row, heads, out_path):
    """원본 grid 전체 + (데이터행에) 결과 열을 붙여 write_only 로 저장.
    PNU 등 원본·결과의 긴 숫자는 문자열이라 지수표기(E+)로 안 깨진다."""
    header_row = start_row - 1   # 결과 열 제목을 놓을 원본 헤더행
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("결과")
    if header_row < 1:           # 원본에 헤더가 없으면(시작=1행) 제목 행을 맨 위에 추가
        ws.append([""] * base_width + list(heads))
    n = len(grid_full)
    for r in range(1, n + 1):
        base = list(grid_full[r - 1])
        if len(base) < base_width:
            base += [""] * (base_width - len(base))
        else:
            base = base[:base_width]
        if r == header_row:
            extra = list(heads)
        elif r >= start_row and r in result_by_row:
            extra = result_by_row[r]
        else:
            extra = [None] * len(heads)
        ws.append(base + extra)
    wb.save(out_path)


# ---------- grid 헬퍼 + 주소 조립 ----------
def n_rows(grid):
    return len(grid)


def n_cols(grid):
    return max((len(r) for r in grid), default=0)


def cell_str(grid, r, c):
    if not c or r < 1 or r > len(grid):
        return ""
    row = grid[r - 1]
    return row[c - 1] if 1 <= c <= len(row) else ""


def build_address(grid, r, sel):
    if sel["kind"] == "full":
        return cell_str(grid, r, sel.get("addr_col"))
    prefix = " ".join(p for p in (cell_str(grid, r, c) for c in sel["admin_cols"]) if p).strip()
    if sel["jibun_kind"] == "cell":
        j = cell_str(grid, r, sel.get("jibun_col"))
    else:
        bon_raw = cell_str(grid, r, sel.get("bon_col"))
        bu_raw = cell_str(grid, r, sel.get("bu_col"))
        is_san = bon_raw.replace(" ", "").startswith("산")
        bon = "".join(ch for ch in bon_raw if ch.isdigit())
        bu = "".join(ch for ch in bu_raw if ch.isdigit())
        core = f"{bon}-{bu}" if (bon and bu and bu != "0") else bon
        j = (f"산 {core}" if is_san else core)
    return f"{prefix} {j}".strip()


def full_addr(grid, r, sel, prefix=""):
    a = build_address(grid, r, sel)
    return f"{prefix} {a}" if (prefix and a) else a


def _row_cell(row, c):
    return row[c - 1] if (c and 1 <= c <= len(row)) else ""


def build_address_row(row, sel):
    if sel["kind"] == "full":
        return _row_cell(row, sel.get("addr_col"))
    prefix = " ".join(p for p in (_row_cell(row, c) for c in sel["admin_cols"]) if p).strip()
    if sel["jibun_kind"] == "cell":
        j = _row_cell(row, sel.get("jibun_col"))
    else:
        bon_raw = _row_cell(row, sel.get("bon_col"))
        bu_raw = _row_cell(row, sel.get("bu_col"))
        is_san = bon_raw.replace(" ", "").startswith("산")
        bon = "".join(ch for ch in bon_raw if ch.isdigit())
        bu = "".join(ch for ch in bu_raw if ch.isdigit())
        core = f"{bon}-{bu}" if (bon and bu and bu != "0") else bon
        j = (f"산 {core}" if is_san else core)
    return f"{prefix} {j}".strip()


def full_addr_row(row, sel, prefix=""):
    a = build_address_row(row, sel)
    return f"{prefix} {a}" if (prefix and a) else a


# ---------- 자동 추정 ----------
def detect_layout(grid):
    max_r = min(n_rows(grid), 300); max_c = min(n_cols(grid), MAX_COLS)
    hits = {}
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = cell_str(grid, r, c)
            if v and any(v.startswith(s) for s in SIDO):
                hits.setdefault(c, []).append((r, v))
    if not hits:
        return {"start_row": 2, "mode": "split", "cols": [1]}
    sido_col = max(hits, key=lambda c: len(hits[c]))
    start_row = min(r for r, _ in hits[sido_col])
    samples = [v for _, v in hits[sido_col]]
    if any((" " in v and any(ch.isdigit() for ch in v)) for v in samples):
        return {"start_row": start_row, "mode": "single", "cols": [sido_col]}
    jibun_re = re.compile(r"^산?\d+(-\d+)?$")
    cols = [sido_col]; c = sido_col + 1
    while c <= max_c and len(cols) < 6:
        vals = [cell_str(grid, rr, c) for rr in range(start_row, min(start_row + 8, max_r + 1))]
        vals = [v for v in vals if v]
        if not vals:
            break
        cols.append(c)
        if sum(1 for v in vals if jibun_re.match(v.replace(" ", ""))) >= max(1, len(vals) // 2):
            break
        c += 1
    return {"start_row": start_row, "mode": "split", "cols": cols}


# ---------- 시트 선택창 ----------
def choose_sheet(parent, sheets):
    dlg = tk.Toplevel(parent); dlg.title("시트 선택"); dlg.configure(bg=CARD)
    dlg.grab_set(); dlg.resizable(False, False)
    tk.Label(dlg, text="변환할 시트를 고르세요", bg=CARD, fg=INK,
             font=(UI_FONT, 14, "bold")).pack(padx=26, pady=(20, 4))
    tk.Label(dlg, text=f"엑셀에 시트가 {len(sheets)}개 있습니다.", bg=CARD, fg=MUTED,
             font=(UI_FONT, 10)).pack(padx=26, pady=(0, 10))
    var = tk.StringVar(value=sheets[0])
    ttk.Combobox(dlg, values=sheets, textvariable=var, width=32,
                 state="readonly").pack(padx=26, pady=6)
    res = {"v": None}

    def ok():
        res["v"] = var.get(); dlg.destroy()

    def cancel():
        res["v"] = None; dlg.destroy()

    brow = tk.Frame(dlg, bg=CARD); brow.pack(pady=18)
    tk.Button(brow, text="확인", command=ok, bg=ACCENT, fg="white", activebackground=ACCENT_D,
              activeforeground="white", relief="flat", width=10, cursor="hand2",
              font=(UI_FONT, 11, "bold")).pack(side="left", padx=8)
    tk.Button(brow, text="취소", command=cancel, bg="#aab8c4", fg="white", activebackground="#94a4b2",
              activeforeground="white", relief="flat", width=8, cursor="hand2",
              font=(UI_FONT, 11)).pack(side="left", padx=8)
    parent.wait_window(dlg)
    return res["v"]


# ---------- CSV 인코딩 선택창 ----------
class EncodingDialog(tk.Toplevel):
    def __init__(self, parent, raw):
        super().__init__(parent)
        self.title("CSV 인코딩 선택"); self.configure(bg=CARD)
        self.raw = raw
        self.result = "__CANCEL__"
        self.resizable(False, False); self.grab_set()

        tk.Label(self, text="CSV 인코딩을 골라 주세요", bg=CARD, fg=INK,
                 font=(UI_FONT, 15, "bold")).pack(padx=24, pady=(18, 2))
        tk.Label(self, text="아래 미리보기에서 한글이 깨지지 않는 걸 고르면 됩니다 (보통 자동 감지로 충분).",
                 bg=CARD, fg=MUTED, font=(UI_FONT, 10)).pack(padx=24, pady=(0, 10))

        self.enc_var = tk.StringVar(value="자동 감지")
        erow = tk.Frame(self, bg=CARD); erow.pack(padx=24, pady=2)
        for label, _enc in CSV_ENCODINGS:
            tk.Radiobutton(erow, text=label, value=label, variable=self.enc_var,
                           command=self.refresh, bg=CARD, fg=INK, selectcolor=SOFT,
                           activebackground=CARD, font=(UI_FONT, 10),
                           anchor="w").pack(anchor="w")

        tk.Label(self, text="미리보기 (처음 몇 줄)", bg=CARD, fg=ACCENT_D,
                 font=(UI_FONT, 10, "bold")).pack(padx=24, pady=(10, 2))
        self.preview = tk.Label(self, text="", bg=SOFT, fg=INK, justify="left", anchor="nw",
                                width=62, height=5, font=(UI_FONT, 10), padx=12, pady=8)
        self.preview.pack(padx=24, pady=4)

        brow = tk.Frame(self, bg=CARD); brow.pack(pady=16)
        tk.Button(brow, text="확인", command=self.ok, bg=ACCENT, fg="white",
                  activebackground=ACCENT_D, activeforeground="white",
                  font=(UI_FONT, 12, "bold"), relief="flat", width=12, cursor="hand2").pack(side="left", padx=8)
        tk.Button(brow, text="취소", command=self.cancel, bg="#aab8c4", fg="white",
                  activebackground="#94a4b2", activeforeground="white",
                  font=(UI_FONT, 12), relief="flat", width=8, cursor="hand2").pack(side="left", padx=8)

        self.refresh()

    def chosen_enc(self):
        label = self.enc_var.get()
        for lab, enc in CSV_ENCODINGS:
            if lab == label:
                return enc
        return None

    def refresh(self):
        try:
            text, used = decode_csv(self.raw, self.chosen_enc())
        except Exception:
            self.preview.config(text="(이 인코딩으로는 읽을 수 없습니다. 다른 것을 골라 주세요.)")
            return
        lines = [ln for ln in text.splitlines() if ln.strip()][:4]
        body = "\n".join(ln[:60] for ln in lines) or "(내용이 비어 있습니다)"
        self.preview.config(text=f"[{used}]\n{body}")

    def ok(self):
        try:
            decode_csv(self.raw, self.chosen_enc())
        except Exception:
            messagebox.showwarning("인코딩", "이 인코딩으로는 파일을 읽을 수 없습니다.", parent=self); return
        self.result = self.chosen_enc(); self.destroy()

    def cancel(self):
        self.result = "__CANCEL__"; self.destroy()


# ---------- 주소 열 확인창 (파일 미리보기 표 포함) ----------
class ColumnDialog(tk.Toplevel):
    def __init__(self, parent, grid, detected):
        super().__init__(parent)
        self.title("주소 열 확인"); self.configure(bg=CARD)
        self.grid_data = grid; self.result = None
        self._ready = False
        self.resizable(False, False); self.grab_set()

        max_c = min(n_cols(grid), MAX_COLS)
        self.opts = ["(없음)"]; self.opt_to_idx = {"(없음)": None}
        for c in range(1, max_c + 1):
            sample = ""
            for rr in range(detected["start_row"], min(detected["start_row"] + 15, n_rows(grid) + 1)):
                s = cell_str(grid, rr, c)
                if s:
                    sample = s[:14]; break
            label = f"{idx_to_col(c)} : {sample}" if sample else f"{idx_to_col(c)} :"
            self.opts.append(label); self.opt_to_idx[label] = c
        self.idx_to_opt = {v: k for k, v in self.opt_to_idx.items()}

        tk.Label(self, text="주소 인식 결과를 확인해 주세요", bg=CARD, fg=INK,
                 font=(UI_FONT, 15, "bold")).grid(row=0, column=0, columnspan=4, pady=(18, 2), padx=24)
        tk.Label(self, text="아래 표에서 내 파일이 어떻게 생겼는지 보고, 주소·본번·부번 열을 지정하세요",
                 bg=CARD, fg=MUTED, font=(UI_FONT, 10)).grid(row=1, column=0, columnspan=4, pady=(0, 8))

        # --- 파일 미리보기 표 (앞 8행 × 앞 10열, A·B·C… 열머리) ---
        tvf = tk.Frame(self, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        tvf.grid(row=2, column=0, columnspan=4, padx=24, pady=(0, 10), sticky="we")
        tk.Label(tvf, text="📄 내 파일 미리보기 (앞부분)", bg=CARD, fg=ACCENT_D,
                 font=(UI_FONT, 10, "bold")).pack(anchor="w", padx=8, pady=(6, 2))
        pv_c = min(n_cols(grid), 10); pv_r = min(n_rows(grid), 8)
        cols = ["행"] + [idx_to_col(c) for c in range(1, pv_c + 1)]
        style = ttk.Style()
        try:
            style.configure("PV.Treeview", font=(UI_FONT, 9), rowheight=22,
                            fieldbackground="white", background="white")
            style.configure("PV.Treeview.Heading", font=(UI_FONT, 9, "bold"))
        except Exception:
            pass
        tv = ttk.Treeview(tvf, columns=cols, show="headings", height=pv_r, style="PV.Treeview")
        tv.heading("행", text="행"); tv.column("행", width=34, anchor="center", stretch=False)
        for c in range(1, pv_c + 1):
            key = idx_to_col(c)
            tv.heading(key, text=key)
            tv.column(key, width=92, anchor="w", stretch=False)
        for rr in range(1, pv_r + 1):
            tv.insert("", "end", values=[rr] + [cell_str(grid, rr, c) for c in range(1, pv_c + 1)])
        xsb = ttk.Scrollbar(tvf, orient="horizontal", command=tv.xview)
        tv.configure(xscrollcommand=xsb.set)
        tv.pack(fill="x", padx=8)
        xsb.pack(fill="x", padx=8, pady=(0, 6))

        row = tk.Frame(self, bg=CARD); row.grid(row=3, column=0, columnspan=4, pady=4)
        tk.Label(row, text="데이터 시작 행:", bg=CARD, fg=INK, font=(UI_FONT, 11)).pack(side="left")
        self.start_var = tk.StringVar(value=str(detected["start_row"]))
        self.start_var.trace_add("write", lambda *a: self.refresh())
        tk.Spinbox(row, from_=1, to=999999, width=8, textvariable=self.start_var,
                   font=(UI_FONT, 11)).pack(side="left", padx=8)

        self.mode_var = tk.StringVar(value="full" if detected["mode"] == "single" else "split")
        mrow = tk.Frame(self, bg=CARD); mrow.grid(row=4, column=0, columnspan=4, pady=8)
        for val, txt in [("full", "한 칸에 전체주소"), ("split", "여러 칸으로 쪼갬")]:
            tk.Radiobutton(mrow, text=txt, value=val, variable=self.mode_var, command=self.on_mode,
                           bg=CARD, fg=INK, selectcolor=SOFT, activebackground=CARD,
                           font=(UI_FONT, 11)).pack(side="left", padx=12)

        self.full_frame = tk.Frame(self, bg=CARD)
        tk.Label(self.full_frame, text="주소 열:", bg=CARD, fg=INK, font=(UI_FONT, 11)).pack(side="left")
        self.full_cb = ttk.Combobox(self.full_frame, values=self.opts, width=26, state="readonly")
        self.full_cb.pack(side="left", padx=8)
        self.full_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        self.split_frame = tk.Frame(self, bg=CARD)
        tk.Label(self.split_frame,
                 text="주소 구성 열을 큰 단위 → 작은 단위 순서로 선택\n(예: 시도·시군구·읍면동·리 등 — 있는 것만, 무엇이든)",
                 bg=CARD, fg=INK, justify="center", font=(UI_FONT, 10)).grid(row=0, column=0, columnspan=3, pady=(0, 4))
        self.admin_cbs = []
        for i in range(5):
            cb = ttk.Combobox(self.split_frame, values=self.opts, width=20, state="readonly")
            cb.grid(row=1 + i // 3, column=i % 3, padx=4, pady=3)
            cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())
            self.admin_cbs.append(cb)

        self.jibun_kind = tk.StringVar(value="cell")
        jrow = tk.Frame(self.split_frame, bg=CARD); jrow.grid(row=3, column=0, columnspan=3, pady=(10, 2))
        tk.Label(jrow, text="지번 형태:", bg=CARD, fg=ACCENT_D, font=(UI_FONT, 10, "bold")).pack(side="left", padx=(0, 6))
        for val, txt in [("cell", "한 칸 (71-2)"), ("bonbu", "본번·부번 분리")]:
            tk.Radiobutton(jrow, text=txt, value=val, variable=self.jibun_kind, command=self.on_jibun,
                           bg=CARD, fg=INK, selectcolor=SOFT, activebackground=CARD,
                           font=(UI_FONT, 10)).pack(side="left", padx=6)

        self.jcell_frame = tk.Frame(self.split_frame, bg=CARD)
        tk.Label(self.jcell_frame, text="지번 열:", bg=CARD, fg=INK, font=(UI_FONT, 10)).pack(side="left")
        self.jibun_cb = ttk.Combobox(self.jcell_frame, values=self.opts, width=20, state="readonly")
        self.jibun_cb.pack(side="left", padx=6)
        self.jibun_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        self.jbonbu_frame = tk.Frame(self.split_frame, bg=CARD)
        tk.Label(self.jbonbu_frame, text="본번 열:", bg=CARD, fg=INK, font=(UI_FONT, 10)).pack(side="left")
        self.bon_cb = ttk.Combobox(self.jbonbu_frame, values=self.opts, width=16, state="readonly")
        self.bon_cb.pack(side="left", padx=6)
        tk.Label(self.jbonbu_frame, text="부번 열:", bg=CARD, fg=INK, font=(UI_FONT, 10)).pack(side="left")
        self.bu_cb = ttk.Combobox(self.jbonbu_frame, values=self.opts, width=16, state="readonly")
        self.bu_cb.pack(side="left", padx=6)
        self.bon_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())
        self.bu_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        self.jcell_frame.grid(row=4, column=0, columnspan=3, pady=4)
        self.jbonbu_frame.grid(row=5, column=0, columnspan=3, pady=4)
        self.full_frame.grid(row=5, column=0, columnspan=4, pady=8)
        self.split_frame.grid(row=6, column=0, columnspan=4, pady=8, padx=24)

        prow = tk.Frame(self, bg=CARD); prow.grid(row=7, column=0, columnspan=4, pady=(6, 0))
        tk.Label(prow, text="앞에 공통으로 붙일 내용(선택):", bg=CARD, fg=INK,
                 font=(UI_FONT, 10)).pack(side="left")
        self.prefix_var = tk.StringVar(value="")
        self.prefix_var.trace_add("write", lambda *a: self.refresh())
        e = tk.Entry(prow, textvariable=self.prefix_var, width=24, font=(UI_FONT, 10))
        e.pack(side="left", padx=8)
        tk.Label(self, text="예: 전북특별자치도 — 데이터에 시도가 빠져 있을 때만 (지역이 섞였으면 비워 두기)",
                 bg=CARD, fg=MUTED, font=(UI_FONT, 9)).grid(row=8, column=0, columnspan=4)

        tk.Label(self, text="미리보기 (실제 변환에 쓰일 주소)", bg=CARD, fg=ACCENT_D,
                 font=(UI_FONT, 10, "bold")).grid(row=9, column=0, columnspan=4, pady=(8, 2))
        self.preview = tk.Label(self, text="", bg=SOFT, fg=INK, justify="left",
                                anchor="w", width=58, height=4, font=(UI_FONT, 10), padx=12)
        self.preview.grid(row=10, column=0, columnspan=4, padx=24, pady=4)

        brow = tk.Frame(self, bg=CARD); brow.grid(row=11, column=0, columnspan=4, pady=16)
        tk.Button(brow, text="변환 시작", command=self.ok, bg=ACCENT, fg="white",
                  activebackground=ACCENT_D, activeforeground="white",
                  font=(UI_FONT, 12, "bold"), relief="flat", width=14, height=1,
                  cursor="hand2").pack(side="left", padx=8)
        tk.Button(brow, text="취소", command=self.cancel, bg="#aab8c4", fg="white",
                  activebackground="#94a4b2", activeforeground="white",
                  font=(UI_FONT, 12), relief="flat", width=8, cursor="hand2").pack(side="left", padx=8)

        if detected["mode"] == "single":
            self.full_cb.set(self.idx_to_opt.get(detected["cols"][0], "(없음)"))
        else:
            admin = detected["cols"][:-1]; jibun = detected["cols"][-1]
            for i, c in enumerate(admin[:5]):
                self.admin_cbs[i].set(self.idx_to_opt.get(c, "(없음)"))
            self.jibun_cb.set(self.idx_to_opt.get(jibun, "(없음)"))
        self._ready = True
        self.on_mode(); self.on_jibun()

    def on_mode(self):
        if self.mode_var.get() == "full":
            self.split_frame.grid_remove(); self.full_frame.grid()
        else:
            self.full_frame.grid_remove(); self.split_frame.grid()
        self.refresh()

    def on_jibun(self):
        if self.jibun_kind.get() == "cell":
            self.jbonbu_frame.grid_remove(); self.jcell_frame.grid()
        else:
            self.jcell_frame.grid_remove(); self.jbonbu_frame.grid()
        self.refresh()

    def idx(self, cb):
        return self.opt_to_idx.get(cb.get())

    def current_sel(self):
        try:
            start = int(self.start_var.get())
        except ValueError:
            start = 1
        prefix = self.prefix_var.get().strip()
        if self.mode_var.get() == "full":
            return {"start_row": start, "kind": "full", "addr_col": self.idx(self.full_cb),
                    "prefix": prefix}
        sel = {"start_row": start, "kind": "split", "prefix": prefix,
               "admin_cols": [self.idx(cb) for cb in self.admin_cbs if self.idx(cb)],
               "jibun_kind": self.jibun_kind.get()}
        if sel["jibun_kind"] == "cell":
            sel["jibun_col"] = self.idx(self.jibun_cb)
        else:
            sel["bon_col"] = self.idx(self.bon_cb); sel["bu_col"] = self.idx(self.bu_cb)
        return sel

    def valid(self, sel):
        if sel["kind"] == "full":
            return bool(sel.get("addr_col"))
        if not sel["admin_cols"]:
            return False
        if sel["jibun_kind"] == "cell":
            return bool(sel.get("jibun_col"))
        return bool(sel.get("bon_col"))

    def refresh(self):
        if not self._ready:
            return
        sel = self.current_sel()
        if not self.valid(sel):
            self.preview.config(text="(주소 열을 선택해 주세요)"); return
        lines, r = [], sel["start_row"]
        while r <= n_rows(self.grid_data) and len(lines) < 3:
            a = full_addr(self.grid_data, r, sel, sel.get("prefix", ""))
            if a and any(ch.isdigit() for ch in a):
                lines.append(f" · {a}")
            r += 1
        self.preview.config(text="\n".join(lines) if lines else "(미리볼 주소가 없습니다)")

    def ok(self):
        sel = self.current_sel()
        if not self.valid(sel):
            messagebox.showwarning("확인", "주소 열을 선택해 주세요.", parent=self); return
        self.result = sel; self.destroy()

    def cancel(self):
        self.result = None; self.destroy()


# ---------- 메인 ----------
class App:
    def __init__(self, root):
        global UI_FONT
        self.root = root; self.tab = 1
        self._valid_key = None
        UI_FONT = _pick_font(root)
        root.title("GIS 주소 변환기"); root.geometry("720x880"); root.configure(bg=BG)
        try:
            root.iconbitmap(resource_path("icon.ico"))
        except Exception:
            pass

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("A.Horizontal.TProgressbar", troughcolor="#dfe8f0",
                        background=ACCENT, bordercolor="#dfe8f0",
                        lightcolor=ACCENT, darkcolor=ACCENT)
        style.configure("TCombobox", fieldbackground="white", background="white")

        header = tk.Frame(root, bg=CARD, height=88); header.pack(fill="x"); header.pack_propagate(False)
        tk.Label(header, text="GIS 주소 변환기", bg=CARD, fg=INK,
                 font=(UI_FONT, 20, "bold")).pack(anchor="w", padx=26, pady=(16, 0))
        tk.Label(header, text="VWorld 기반 · 대용량 병렬 변환 (PNU·좌표·지도 레이어)", bg=CARD, fg=MUTED,
                 font=(UI_FONT, 10)).pack(anchor="w", padx=26)
        tk.Frame(root, bg=ACCENT, height=3).pack(fill="x")

        tabbar = tk.Frame(root, bg=BG); tabbar.pack(fill="x", padx=18, pady=(14, 0))
        self.tab1_btn = tk.Button(tabbar, text="①  주소 → PNU", command=lambda: self.switch(1),
                                  font=(UI_FONT, 11, "bold"), relief="flat", bd=0, height=2, cursor="hand2")
        self.tab2_btn = tk.Button(tabbar, text="②  주소 → 좌표", command=lambda: self.switch(2),
                                  font=(UI_FONT, 11, "bold"), relief="flat", bd=0, height=2, cursor="hand2")
        self.tab3_btn = tk.Button(tabbar, text="③  QGIS 레이어", command=lambda: self.switch(3),
                                  font=(UI_FONT, 11, "bold"), relief="flat", bd=0, height=2, cursor="hand2")
        self.tab1_btn.pack(side="left", expand=True, fill="x", padx=(0, 3))
        self.tab2_btn.pack(side="left", expand=True, fill="x", padx=3)
        self.tab3_btn.pack(side="left", expand=True, fill="x", padx=(3, 0))

        card = tk.Frame(root, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        card.pack(fill="x", padx=18, pady=14)
        self.card = card

        # 🔑 인증키 입력 — 앱에서 바로 붙여넣고 저장 (파일 편집 불필요)
        keyf = tk.Frame(card, bg=CARD); keyf.pack(fill="x", padx=22, pady=(16, 4))
        tk.Label(keyf, text="🔑 VWorld 인증키", bg=CARD, fg=INK,
                 font=(UI_FONT, 11, "bold")).pack(anchor="w")
        krow = tk.Frame(keyf, bg=CARD); krow.pack(fill="x", pady=(4, 0))
        self.key_var = tk.StringVar(value="")
        self.key_entry = tk.Entry(krow, textvariable=self.key_var, show="●",
                                  font=(UI_FONT, 11), relief="solid", bd=1)
        self.key_entry.pack(side="left", fill="x", expand=True, ipady=3)
        self.show_key = tk.BooleanVar(value=False)
        tk.Checkbutton(krow, text="보기", variable=self.show_key, command=self._toggle_key,
                       bg=CARD, fg=MUTED, activebackground=CARD, selectcolor=SOFT,
                       font=(UI_FONT, 9)).pack(side="left", padx=6)
        tk.Button(krow, text="확인·저장", command=self.verify_and_save, bg=ACCENT, fg="white",
                  activebackground=ACCENT_D, activeforeground="white", relief="flat",
                  font=(UI_FONT, 10, "bold"), cursor="hand2").pack(side="left", padx=(4, 0))
        self.key_status = tk.Label(keyf, text="", bg=CARD, fg=MUTED, font=(UI_FONT, 10), anchor="w")
        self.key_status.pack(anchor="w", pady=(5, 0))
        khelp = tk.Frame(keyf, bg=CARD); khelp.pack(anchor="w")
        tk.Label(khelp, text="키가 없으면 무료 발급 후 여기에 붙여넣고 [확인·저장] →",
                 bg=CARD, fg=MUTED, font=(UI_FONT, 9)).pack(side="left")
        tk.Button(khelp, text="vworld.kr 열기", relief="flat", bg=CARD, fg=ACCENT_D,
                  activebackground=CARD, cursor="hand2", font=(UI_FONT, 9, "underline"),
                  command=lambda: webbrowser.open("https://www.vworld.kr")).pack(side="left")
        tk.Frame(card, bg=LINE, height=1).pack(fill="x", padx=22, pady=(8, 2))

        self.desc = tk.Label(card, text="", bg=CARD, fg=INK, justify="left", anchor="w",
                             font=(UI_FONT, 11), wraplength=620)
        self.desc.pack(fill="x", padx=22, pady=(10, 8))

        self.opt_area = tk.Frame(card, bg=CARD); self.opt_area.pack(fill="x", padx=22)
        self.jiga_var = tk.BooleanVar(value=False)
        self.opt_pnu = tk.Frame(self.opt_area, bg=CARD)
        tk.Checkbutton(self.opt_pnu, text="공시지가(원/㎡)·기준연월도 함께 조회합니다  (조회량 2배 — 조금 느려집니다)",
                       variable=self.jiga_var, bg=CARD, fg=INK, activebackground=CARD,
                       selectcolor=SOFT, font=(UI_FONT, 11), cursor="hand2").pack(anchor="w")
        self.opt_geo = tk.Frame(self.opt_area, bg=CARD)
        tk.Label(self.opt_geo, text="좌표계:", bg=CARD, fg=INK, font=(UI_FONT, 11)).pack(side="left")
        self.crs_cb = ttk.Combobox(self.opt_geo, values=list(CRS_OPTIONS), width=24, state="readonly")
        self.crs_cb.current(0); self.crs_cb.pack(side="left", padx=8)

        self.pt_var = tk.BooleanVar(value=True); self.pg_var = tk.BooleanVar(value=True)
        self.opt_layer = tk.Frame(self.opt_area, bg=CARD)
        tk.Checkbutton(self.opt_layer, text="포인트 레이어 — 주소를 점으로 (빠릅니다)", variable=self.pt_var,
                       bg=CARD, fg=INK, activebackground=CARD, selectcolor=SOFT,
                       font=(UI_FONT, 11), cursor="hand2").pack(anchor="w")
        tk.Checkbutton(self.opt_layer, text="필지 경계 레이어 — 실제 땅 모양·공시지가 포함 (느릴 수 있습니다)",
                       variable=self.pg_var, bg=CARD, fg=INK, activebackground=CARD, selectcolor=SOFT,
                       font=(UI_FONT, 11), cursor="hand2").pack(anchor="w")

        self.btn = tk.Button(card, text=PICK_TEXT, command=self.pick,
                             font=(UI_FONT, 13, "bold"), bg=ACCENT, fg="white",
                             activebackground=ACCENT_D, activeforeground="white",
                             relief="flat", height=2, cursor="hand2")
        self.btn.pack(fill="x", padx=22, pady=(16, 8))

        prow = tk.Frame(card, bg=CARD); prow.pack(fill="x", padx=22, pady=(0, 18))
        self.progress = ttk.Progressbar(prow, style="A.Horizontal.TProgressbar",
                                        mode="determinate", maximum=100)
        self.progress.pack(side="left", fill="x", expand=True)
        self.pct_lbl = tk.Label(prow, text="0%", bg=CARD, fg=MUTED, width=18,
                                font=(UI_FONT, 10))
        self.pct_lbl.pack(side="left", padx=(10, 0))

        self.log = scrolledtext.ScrolledText(root, height=10, font=("Consolas", 10),
                                             bg=LOGBG, fg=INK, relief="flat",
                                             highlightbackground=LINE, highlightthickness=1)
        self.log.pack(fill="both", expand=True, padx=18, pady=(0, 16))

        self.cfg = load_config()
        self.write(f"UI 글꼴: {UI_FONT}")
        saved = self.cfg.get("API_KEY", "")
        if saved and "여기에" not in saved:
            self.key_var.set(saved.strip())
            self._set_key_status("saved")
            self.write("✔ 저장된 인증키를 불러왔어요. 기능을 고르고 파일을 올려 주세요.")
        else:
            self._set_key_status("empty")
            self.write("👉 먼저 위쪽 '인증키' 칸에 VWorld 인증키를 붙여넣고 [확인·저장]을 눌러 주세요.")
        self.write(f"※ 동시 처리 {self._workers()}개(config.txt WORKERS). 통신실패가 많으면 값을 줄여 보세요.")
        self.switch(1)

        for w in (self.root, self.card, self.btn, self.log, self.desc, self.opt_area):
            try:
                w.drop_target_register(DND_FILES)
                w.dnd_bind("<<Drop>>", self.on_drop)
            except Exception as e:
                self.write(f"(드래그앤드롭 일부 비활성화: {e})")

    def _workers(self):
        try:
            w = int(self.cfg.get("WORKERS", WORKERS_DEFAULT))
        except (ValueError, TypeError):
            w = WORKERS_DEFAULT
        return max(1, min(POOL_SIZE, w))

    def switch(self, n):
        self.tab = n
        for k, b in {1: self.tab1_btn, 2: self.tab2_btn, 3: self.tab3_btn}.items():
            if k == n:
                b.config(bg=ACCENT, fg="white")
            else:
                b.config(bg=TAB_OFF_BG, fg=MUTED)
        self.opt_pnu.pack_forget(); self.opt_geo.pack_forget(); self.opt_layer.pack_forget()
        if n == 1:
            self.desc.config(text="주소를 19자리 PNU로 변환합니다. 원본 열은 그대로 두고 결과 열을 오른쪽에 덧붙입니다.\n결과: 입력주소 · PNU · 정제주소 · 본번 · 부번 (선택 시 공시지가·기준연월)")
            self.opt_pnu.pack(anchor="w")
        elif n == 2:
            self.desc.config(text="주소를 지도 좌표로 변환합니다. 원본 열은 그대로 두고 결과 열을 오른쪽에 덧붙입니다.\n결과: 입력주소 · 선택 좌표계의 X · Y · 정제주소")
            self.opt_geo.pack(anchor="w")
        else:
            self.desc.config(text="주소를 QGIS에서 바로 열리는 지도 레이어(GeoJSON)로 만듭니다.\nQGIS 창에 파일을 끌어다 놓으면 점/필지로 표시됩니다.")
            self.opt_layer.pack(anchor="w")

    def write(self, msg):
        self.log.insert("end", msg + "\n"); self.log.see("end"); self.root.update_idletasks()

    def set_progress(self, done, total):
        total = max(1, total)
        self.progress["maximum"] = total
        self.progress["value"] = done
        self.pct_lbl.config(text=f"{int(done / total * 100)}%  ({done:,}/{total:,})")
        self.root.update_idletasks()

    def _toggle_key(self):
        self.key_entry.config(show="" if self.show_key.get() else "●")

    def _set_key_status(self, kind, msg=""):
        styles = {
            "ok": ("✅ 인증키 정상 — 저장했어요. 바로 쓸 수 있습니다.", "#2e7d32"),
            "saved": ("✅ 저장된 인증키를 불러왔어요.", "#2e7d32"),
            "checking": ("인증키 확인 중…", MUTED),
            "invalid": ("❌ 인증키가 올바르지 않아요. 다시 붙여넣어 확인해 주세요.", "#c0392b"),
            "network": ("⚠️ 지금 VWorld 연결이 안 돼 확인 못 했어요. 잠시 뒤 다시 눌러 주세요.", "#b9770e"),
            "empty": ("인증키를 붙여넣고 [확인·저장]을 눌러 주세요.", MUTED),
        }
        text, color = styles.get(kind, ("", MUTED))
        self.key_status.config(text=msg or text, fg=color)

    def current_key(self):
        return self.key_var.get().strip()

    def verify_and_save(self):
        key = self.current_key()
        if not key:
            self._set_key_status("empty"); return
        self._set_key_status("checking"); self.root.update_idletasks()
        status, _detail = validate_key(key)
        if status == "invalid":
            self._valid_key = None; self._set_key_status("invalid"); return
        if status == "network":
            self._set_key_status("network"); return
        set_config_value("API_KEY", key)
        self.cfg = load_config()
        self._valid_key = key
        self._set_key_status("ok")
        self.write("✔ 인증키 저장 완료.")

    def check_api_key(self):
        """인증키 칸의 값을 확인·저장하고 유효하면 반환, 아니면 안내 후 None."""
        key = self.current_key()
        if not key:   # 칸이 비었으면 저장돼 있던 키라도 불러온다
            self.cfg = load_config()
            saved = self.cfg.get("API_KEY", "")
            if saved and "여기에" not in saved:
                key = saved.strip(); self.key_var.set(key)
        if not key or "여기에" in key:
            messagebox.showinfo("인증키 필요",
                "먼저 위쪽 '인증키' 칸에 VWorld 인증키를 붙여넣고 [확인·저장]을 눌러 주세요.")
            try:
                self.key_entry.focus_set()
            except Exception:
                pass
            self._set_key_status("empty"); return None
        if self._valid_key == key:
            return key
        self.write("인증키 확인 중…")
        status, _detail = validate_key(key)
        if status == "invalid":
            messagebox.showerror("인증키 오류",
                "VWorld 인증키가 올바르지 않습니다. 위쪽 칸에서 다시 확인해 주세요.")
            self._set_key_status("invalid"); return None
        if status == "network":
            if not messagebox.askokcancel("연결 확인",
                    "지금 VWorld에 연결이 안 돼 인증키를 확인하지 못했습니다.\n그래도 진행할까요?"):
                return None
            self.write("⚠ 인증키 확인 못함(연결). 그대로 진행합니다.")
            return key
        set_config_value("API_KEY", key)
        self.cfg = load_config()
        self._valid_key = key
        self._set_key_status("ok")
        self.write("✔ 인증키 정상")
        return key

    def pick(self):
        if self.check_api_key() is None:
            return
        path = filedialog.askopenfilename(
            title="파일 선택",
            filetypes=[("주소 파일", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if path:
            self.start_with_file(path)

    def on_drop(self, event):
        if str(self.btn["state"]) == "disabled":
            return
        try:
            dropped = list(self.root.tk.splitlist(event.data))
        except Exception:
            dropped = [event.data]
        paths = [p for p in dropped if p.lower().endswith((".xlsx", ".csv"))]
        if not paths:
            messagebox.showwarning("파일 형식", "엑셀(.xlsx) 또는 CSV(.csv) 파일을 끌어다 놓아 주세요.")
            return
        self.root.after(0, lambda: self.start_with_file(paths[0]))

    def start_with_file(self, path):
        api_key = self.check_api_key()
        if api_key is None:
            return
        ext = path.lower().rsplit(".", 1)[-1]
        raw = None; enc = None; sheet = None

        if ext == "csv":
            try:
                with open(path, "rb") as f:
                    raw = f.read()
            except Exception as e:
                messagebox.showerror("오류", f"파일을 읽을 수 없습니다:\n{e}"); return
            dlg = EncodingDialog(self.root, raw)
            self.root.wait_window(dlg)
            if dlg.result == "__CANCEL__":
                self.write("취소되었습니다."); return
            enc = dlg.result
            try:
                head, enc = read_head(path, "csv", None, enc, raw=raw)
            except Exception as e:
                messagebox.showerror("오류", f"CSV를 읽을 수 없습니다:\n{e}"); return
            self.write(f"CSV 인코딩: {enc}")
        else:
            try:
                sheets = xlsx_sheet_names(path)
            except Exception as e:
                messagebox.showerror("오류", f"엑셀을 열 수 없습니다:\n{e}"); return
            sheet = sheets[0] if sheets else None
            if len(sheets) > 1:
                sheet = choose_sheet(self.root, sheets)
                if sheet is None:
                    self.write("취소되었습니다."); return
            try:
                head, _ = read_head(path, "xlsx", sheet, None)
            except Exception as e:
                messagebox.showerror("오류", f"시트를 열 수 없습니다:\n{e}"); return

        if n_rows(head) == 0:
            messagebox.showwarning("빈 파일", "내용이 없는 파일입니다."); return
        detected = detect_layout(head)
        dlg = ColumnDialog(self.root, head, detected)
        self.root.wait_window(dlg)
        if not dlg.result:
            self.write("취소되었습니다."); return
        sel = dlg.result
        opts = {"jiga": self.jiga_var.get(), "crs_label": self.crs_cb.get(),
                "pt": self.pt_var.get(), "pg": self.pg_var.get()}
        self.set_progress(0, 1)
        self.btn.config(state="disabled", text="처리 중...")
        threading.Thread(target=self.run,
                         args=(path, ext, sheet, enc, raw, sel, self.tab, opts),
                         daemon=True).start()

    def run(self, path, ext, sheet, enc, raw, sel, tab, opts):
        try:
            api_key = self.current_key() or self.cfg.get("API_KEY", "").strip()
            domain = self.cfg.get("DOMAIN", "localhost")
            workers = self._workers()
            start_row = sel["start_row"]; prefix = sel.get("prefix", "")

            if tab == 3:
                if not (opts["pt"] or opts["pg"]):
                    messagebox.showwarning("선택", "포인트/필지 중 하나는 체크해 주세요.")
                    return
                self.run_layers(path, ext, sheet, enc, raw, sel, opts, api_key, domain, workers)
                return

            # ①②: 원본 열 보존 → 전체 grid 를 읽는다
            self.write("파일을 읽는 중…")
            grid_full, truncated = read_full_grid(path, ext, sheet, enc, raw=raw)
            if truncated:
                self.write(f"⚠ 너무 많아 {MAX_ROWS:,}행까지만 처리합니다.")
            base_width = max((len(r) for r in grid_full), default=0)
            total_rows = max(0, len(grid_full) - start_row + 1)

            row_addr = {}
            for r in range(start_row, len(grid_full) + 1):
                row_addr[r] = full_addr_row(grid_full[r - 1], sel, prefix)
            valid_rows = [r for r in range(start_row, len(grid_full) + 1)
                          if row_addr[r] and any(ch.isdigit() for ch in row_addr[r])]
            skip = total_rows - len(valid_rows)
            if not valid_rows:
                self.write("변환할 주소를 찾지 못했습니다.")
                messagebox.showwarning("확인", "주소를 찾지 못했습니다. 열/시작 행을 확인해 주세요.")
                return

            crs = CRS_OPTIONS[opts["crs_label"]][0] if tab == 2 else "EPSG:4326"
            need_parcel = (tab == 1 and opts["jiga"])
            total = len(valid_rows)
            self.write(f"변환 시작: {total:,}건 (동시 {workers}개, 건너뜀 {skip:,})")

            results = {}
            done = 0
            self.set_progress(0, total)
            from concurrent.futures import ThreadPoolExecutor, as_completed
            with ThreadPoolExecutor(max_workers=workers) as ex:
                fut2r = {ex.submit(work, row_addr[r], api_key, crs, need_parcel, domain): r
                         for r in valid_rows}
                for fut in as_completed(fut2r):
                    results[fut2r[fut]] = fut.result()
                    done += 1
                    if done % 100 == 0 or done == total:
                        self.set_progress(done, total)
                    if done % 2000 == 0:
                        self.write(f"  … {done:,}/{total:,}")

            self.save_records(path, tab, sel, opts, grid_full, base_width, row_addr, results, valid_rows, skip)
        except Exception as e:
            self.write(f"\n[오류] {e}")
            messagebox.showerror("오류", str(e))
        finally:
            self.btn.config(state="normal", text=PICK_TEXT)

    def save_records(self, path, tab, sel, opts, grid_full, base_width, row_addr, results, valid_rows, skip):
        start_row = sel["start_row"]
        ok = fail = commfail = 0
        result_by_row = {}
        if tab == 1:
            want_jiga = opts["jiga"]
            heads = (["입력주소", "PNU", "정제주소", "본번", "부번"]
                     + (["공시지가(원/㎡)", "기준연월"] if want_jiga else []) + ["상태"])
            for r in valid_rows:
                item = results.get(r) or {"status": "실패", "pnu": None, "refined": None}
                status = item["status"]; pnu = item["pnu"]
                st_ = "PNU불완전" if (status == "OK" and not (pnu and len(pnu) == 19)) else status
                if status == "OK":
                    ok += 1
                else:
                    fail += 1
                    if str(status).startswith("통신실패"):
                        commfail += 1
                bon, bu = parse_bonbu(pnu)
                vals = [row_addr[r], pnu, item.get("refined"), bon, bu]
                if want_jiga:
                    props = item.get("props") or {}
                    jg = props.get("jiga")
                    vals += [int(jg) if jg and str(jg).isdigit() else None,
                             (f"{props.get('gosi_year','')}.{props.get('gosi_month','')}".strip(".") if props else None)]
                vals += [st_]
                result_by_row[r] = vals
        else:
            xlab, ylab = CRS_OPTIONS[opts["crs_label"]][1], CRS_OPTIONS[opts["crs_label"]][2]
            heads = ["입력주소", xlab, ylab, "정제주소", "상태"]
            for r in valid_rows:
                item = results.get(r) or {"status": "실패", "x": None, "y": None, "refined": None}
                status = item["status"]
                if status == "OK":
                    ok += 1
                else:
                    fail += 1
                    if str(status).startswith("통신실패"):
                        commfail += 1
                result_by_row[r] = [row_addr[r], item.get("x"), item.get("y"), item.get("refined"), status]

        out_path = os.path.splitext(path)[0] + "_결과.xlsx"
        self.write("결과 저장 중…")
        save_with_original(grid_full, base_width, start_row, result_by_row, heads, out_path)
        self.set_progress(1, 1)
        self.write(f"\n✅ 완료 · 성공 {ok:,} / 실패 {fail:,} / 건너뜀 {skip:,}")
        if commfail:
            self.write(f"※ 통신실패 {commfail:,}건 — VWorld 연결이 조여졌을 수 있어요. "
                       f"config.txt 의 WORKERS 를 줄이거나(예: {max(1, self._workers()-1)}) 잠시 뒤 실패분만 다시 돌려 보세요.")
        self.write(f"저장: {out_path}")
        messagebox.showinfo("완료",
            f"성공 {ok:,}건 / 실패 {fail:,}건 / 건너뜀 {skip:,}건\n\n저장되었습니다:\n{out_path}")

    def run_layers(self, path, ext, sheet, enc, raw, sel, opts, api_key, domain, workers):
        start_row = sel["start_row"]; prefix = sel.get("prefix", "")
        self.write("주소를 모으는 중…")
        addrs, truncated = collect_addresses(path, ext, sheet, enc, sel, prefix, start_row, raw=raw)
        if truncated:
            self.write(f"⚠ 너무 많아 {MAX_ROWS:,}건까지만 처리합니다.")
        valid = [a for a in addrs if a and any(ch.isdigit() for ch in a)]
        skip = len(addrs) - len(valid)
        if not valid:
            self.write("변환할 주소를 찾지 못했습니다.")
            messagebox.showwarning("확인", "주소를 찾지 못했습니다."); return

        want_pt, want_pg = opts["pt"], opts["pg"]
        total = len(valid)
        self.write(f"변환 시작: {total:,}건 (동시 {workers}개, 건너뜀 {skip:,})")
        results = [None] * total
        done = 0
        self.set_progress(0, total)
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=workers) as ex:
            fut2pos = {ex.submit(work, a, api_key, "EPSG:4326", want_pg, domain): p
                       for p, a in enumerate(valid)}
            for fut in as_completed(fut2pos):
                results[fut2pos[fut]] = fut.result()
                done += 1
                if done % 100 == 0 or done == total:
                    self.set_progress(done, total)
                if done % 2000 == 0:
                    self.write(f"  … {done:,}/{total:,}")

        pts, pgs = [], []
        ok = fail = 0
        for item in results:
            status = item["status"]; x = item["x"]; y = item["y"]
            pnu = item["pnu"]; refined = item["refined"]; addr = item["addr"]
            if status != "OK" or not x:
                fail += 1; continue
            ok += 1
            bon, bu = parse_bonbu(pnu)
            if want_pt:
                pts.append({"type": "Feature",
                            "geometry": {"type": "Point", "coordinates": [float(x), float(y)]},
                            "properties": {"입력주소": addr, "정제주소": refined,
                                           "PNU": pnu, "본번": bon, "부번": bu}})
            if want_pg:
                geom = item.get("geom"); props = item.get("props") or {}
                if geom:
                    jg = props.get("jiga")
                    pgs.append({"type": "Feature", "geometry": geom,
                                "properties": {"정제주소": refined, "PNU": props.get("pnu") or pnu,
                                               "본번": bon, "부번": bu,
                                               "공시지가": int(jg) if jg and str(jg).isdigit() else None,
                                               "기준연월": f"{props.get('gosi_year','')}.{props.get('gosi_month','')}".strip(".")}})

        base = os.path.splitext(path)[0]; saved = []
        if want_pt and pts:
            p = base + "_포인트.geojson"
            with open(p, "w", encoding="utf-8") as f:
                json.dump({"type": "FeatureCollection", "features": pts}, f, ensure_ascii=False)
            saved.append(p)
        if want_pg and pgs:
            p = base + "_필지.geojson"
            with open(p, "w", encoding="utf-8") as f:
                json.dump({"type": "FeatureCollection", "features": pgs}, f, ensure_ascii=False)
            saved.append(p)

        self.set_progress(1, 1)
        self.write(f"\n✅ 완료 · 성공 {ok:,} / 실패 {fail:,}")
        for s in saved:
            self.write("저장: " + s)
        if saved:
            messagebox.showinfo("완료",
                f"성공 {ok:,}건\nQGIS 창에 파일을 끌어다 놓으면 됩니다.\n\n" + "\n".join(saved))
        else:
            messagebox.showwarning("완료", "저장할 결과가 없습니다. (모두 실패하거나 건너뜀)")


if __name__ == "__main__":
    try:
        root = TkinterDnD.Tk()
    except Exception:
        root = tk.Tk()
    App(root)
    root.mainloop()
