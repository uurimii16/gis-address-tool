# -*- coding: utf-8 -*-
"""
GIS 주소 변환기 (웹 / Streamlit)
 ① 주소 → PNU   ② 주소 → 좌표   ③ QGIS 레이어(GeoJSON)
실행:  streamlit run streamlit_app.py
"""
import io
import re
import csv
import json
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

GEOCODE_URL = "https://api.vworld.kr/req/address"
DATA_URL = "https://api.vworld.kr/req/data"

MAX_WORKERS = 8          # 동시에 물어보는 주소 개수 (병렬 처리)
MAX_ROWS = 100_000       # 안전을 위한 최대 처리 행 수
MAX_COLS = 60            # 읽어들일 최대 열 수
XLSX_SIZE_LIMIT_MB = 20  # 엑셀 업로드 한도 (read_only 로딩이라 크게 잡아도 안전)
CSV_SIZE_LIMIT_MB = 40   # CSV 업로드 한도 (메모리 부담이 더 적음)

SIDO = ["서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시", "대전광역시",
        "울산광역시", "세종특별자치시", "경기도", "강원특별자치도", "강원도", "충청북도",
        "충청남도", "전북특별자치도", "전라북도", "전라남도", "경상북도", "경상남도",
        "제주특별자치도", "충북", "충남", "전북", "전남", "경북", "경남", "경기", "강원", "제주"]

CRS_OPTIONS = {
    "위경도 (EPSG:4326)": ("EPSG:4326", "경도(X)", "위도(Y)"),
    "중부원점TM (EPSG:5186)": ("EPSG:5186", "X(TM)", "Y(TM)"),
}

CSV_ENCODINGS = {
    "자동 감지": None,
    "CP949 / EUC-KR (한글 윈도우·공공데이터 CSV)": "cp949",
    "UTF-8": "utf-8-sig",
}

REQ_TIMEOUT = 10   # 한 번 호출을 기다리는 최대 시간(초) — 넘으면 재시도

# 커넥션 재사용용 세션 (TLS 핸드셰이크 재활용 → 정상일 때 더 빠름).
# 일부 방화벽/WAF는 기본 python-requests UA를 막으므로 UA도 지정한다.
_session = requests.Session()
_session.headers.update({"User-Agent": "gis-address-tool/1.0 (+github.com/uurimii16/gis-address-tool)"})
_adapter = requests.adapters.HTTPAdapter(
    pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS, max_retries=0)
_session.mount("https://", _adapter)
_session.mount("http://", _adapter)

# 회로 차단기: 연결실패가 이만큼 쌓이면 재시도를 멈추고 빨리 실패시킨다.
# (VWorld 연결이 통째로 막힌 환경에서 주소마다 재시도하며 한없이 느려지는 것 방지)
_CIRCUIT_TRIP = 8
_conn_fail = {"n": 0}   # 스레드 간 공유. 대략치라 락 없이 사용(오차 무해).


def _vworld_json(url, params, timeout=REQ_TIMEOUT):
    """VWorld를 호출해 (json, None) 또는 (None, 사람이 읽을 실패사유)를 반환.
    실패사유가 있으면 대개 '일시적'(빈응답·연결끊김·5xx)이라 재시도 가치가 있다."""
    try:
        r = _session.get(url, params=params, timeout=timeout)
    except requests.RequestException as e:
        return None, "연결실패:" + type(e).__name__
    if r.status_code != 200:
        return None, f"HTTP{r.status_code}"
    try:
        return r.json(), None
    except ValueError:
        body = (r.text or "").strip()
        # VWorld는 정상이면 항상 JSON을 준다. 본문이 비었으면 IP 차단/스로틀링 정황.
        return None, ("빈응답(IP제한 의심)" if not body
                      else "비정상응답:" + body[:60].replace("\n", " "))


# ---------- 파일 읽기 (xlsx·csv 모두 문자열 2차원 리스트 grid 로 통일; r,c는 1-based) ----------
def _normalize(row_iter):
    """행 반복자를 문자열 2차원 리스트로 변환. 열은 MAX_COLS, 행은 MAX_ROWS로 상한."""
    grid, truncated = [], False
    for row in row_iter:
        grid.append(["" if v is None else str(v).strip() for v in row[:MAX_COLS]])
        if len(grid) >= MAX_ROWS:
            truncated = True
            break
    return grid, truncated


def xlsx_sheet_names(file_bytes):
    """엑셀의 시트 이름 목록을 반환(가볍게 열고 닫음)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def read_xlsx_grid(file_bytes, sheet=None):
    """엑셀의 지정 시트를 read_only(스트리밍) 모드로 훑어 grid 로 변환. sheet=None이면 첫(활성) 시트."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
        return _normalize(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def read_csv_grid(data_bytes, enc):
    """CSV 바이트를 grid 로 변환. enc=None이면 여러 인코딩을 순서대로 시도(자동 감지)."""
    tries = [enc] if enc else ["utf-8-sig", "cp949", "euc-kr", "latin1"]
    text = used = last = None
    for e in tries:
        try:
            text = data_bytes.decode(e); used = e; break
        except Exception as ex:
            last = ex
    if text is None:
        raise last or ValueError("인코딩을 인식하지 못했습니다.")
    grid, truncated = _normalize(csv.reader(io.StringIO(text)))
    return grid, used, truncated


def n_rows(grid):
    return len(grid)


def n_cols(grid):
    return max((len(r) for r in grid), default=0)


def cell_str(grid, r, c):
    if not c or r < 1 or r > len(grid):
        return ""
    row = grid[r - 1]
    return row[c - 1] if 1 <= c <= len(row) else ""


# ---------- 공통 로직 ----------
def idx_to_col(i):
    s = ""
    while i > 0:
        i, rem = divmod(i - 1, 26)
        s = chr(65 + rem) + s
    return s


def parse_bonbu(pnu):
    if pnu and len(pnu) == 19:
        return int(pnu[11:15]), int(pnu[15:19])
    return None, None


def build_address(grid, r, sel):
    if sel["kind"] == "full":
        return cell_str(grid, r, sel.get("addr_col"))
    prefix = " ".join(p for p in (cell_str(grid, r, c) for c in sel["admin_cols"]) if p).strip()
    if sel["jibun_kind"] == "cell":
        j = cell_str(grid, r, sel.get("jibun_col"))
    else:
        bon_raw = cell_str(grid, r, sel.get("bon_col"))
        bu_raw = cell_str(grid, r, sel.get("bu_col"))
        is_san = bon_raw.replace(" ", "").startswith("산")
        bon = "".join(ch for ch in bon_raw if ch.isdigit())
        bu = "".join(ch for ch in bu_raw if ch.isdigit())
        core = f"{bon}-{bu}" if (bon and bu and bu != "0") else bon
        j = (f"산 {core}" if is_san else core)
    return f"{prefix} {j}".strip()


def full_addr(grid, r, sel, prefix=""):
    """조립한 주소 앞에 공통 접두어(예: 시도명)를 붙인다. 접두어가 없거나 주소가 비면 그대로."""
    a = build_address(grid, r, sel)
    return f"{prefix} {a}" if (prefix and a) else a


def detect_layout(grid):
    max_r = min(n_rows(grid), 300); max_c = min(n_cols(grid), MAX_COLS)
    hits = {}
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = cell_str(grid, r, c)
            if v and any(v.startswith(s) for s in SIDO):
                hits.setdefault(c, []).append((r, v))
    if not hits:
        return {"start_row": 2, "mode": "split", "cols": [1]}
    sido_col = max(hits, key=lambda c: len(hits[c]))
    start_row = min(r for r, _ in hits[sido_col])
    samples = [v for _, v in hits[sido_col]]
    if any((" " in v and any(ch.isdigit() for ch in v)) for v in samples):
        return {"start_row": start_row, "mode": "single", "cols": [sido_col]}
    jibun_re = re.compile(r"^산?\d+(-\d+)?$")
    cols = [sido_col]; c = sido_col + 1
    while c <= max_c and len(cols) < 6:
        vals = [cell_str(grid, rr, c) for rr in range(start_row, min(start_row + 8, max_r + 1))]
        vals = [v for v in vals if v]
        if not vals:
            break
        cols.append(c)
        if sum(1 for v in vals if jibun_re.match(v.replace(" ", ""))) >= max(1, len(vals) // 2):
            break
        c += 1
    return {"start_row": start_row, "mode": "split", "cols": cols}


def geocode(addr, api_key, crs="EPSG:4326", retries=3):
    last = "실패"
    # 연결실패가 이미 많이 쌓였으면 재시도 없이 1번만 시도(빨리 실패)
    attempts = 1 if _conn_fail["n"] >= _CIRCUIT_TRIP else retries
    for attempt in range(attempts):
        j, err = _vworld_json(GEOCODE_URL, {
            "service": "address", "request": "getcoord", "version": "2.0",
            "crs": crs, "address": addr, "type": "PARCEL",
            "format": "json", "key": api_key})
        if err is None:
            _conn_fail["n"] = 0   # 한 번이라도 응답이 오면 회로 복구
            resp = j.get("response", {})
            if resp.get("status") == "OK":
                try:
                    s = resp["refined"]["structure"]
                    pt = resp["result"]["point"]
                    return (s.get("level4LC", ""), pt["x"], pt["y"],
                            resp["refined"].get("text", ""), "OK")
                except (KeyError, TypeError):
                    return None, None, None, None, "주소인식실패"
            # VWorld가 확정 답(에러/미검색)을 준 경우 — 재시도해도 결과 동일하니 즉시 종료
            code = (resp.get("error") or {}).get("code", "")
            if code == "INVALID_KEY":
                return None, None, None, None, "인증키오류"
            return None, None, None, None, "주소인식실패"
        # 여기 도달 = 빈응답·연결끊김·5xx 등 일시적 → 잠깐 쉬고 재시도
        last = err
        if attempt < attempts - 1:
            time.sleep(0.4 * (attempt + 1))
    _conn_fail["n"] += 1
    return None, None, None, None, f"통신실패:{last}"


def get_parcel(x, y, api_key, domain="localhost", retries=3):
    for attempt in range(retries):
        j, err = _vworld_json(DATA_URL, {
            "service": "data", "request": "GetFeature", "data": "LP_PA_CBND_BUBUN",
            "version": "2.0", "geomFilter": f"POINT({x} {y})", "crs": "EPSG:4326",
            "format": "json", "size": "1", "domain": domain, "key": api_key})
        if err is None:
            resp = j.get("response", {})
            if resp.get("status") != "OK":
                return None, None
            try:
                feat = resp["result"]["featureCollection"]["features"][0]
                return feat.get("geometry"), feat.get("properties", {})
            except (KeyError, IndexError, TypeError):
                return None, None
        if attempt < retries - 1:
            time.sleep(0.4 * (attempt + 1))
    return None, None


def build_options(grid, start_row):
    opts = ["(없음)"]; o2i = {"(없음)": None}
    for c in range(1, min(n_cols(grid), MAX_COLS) + 1):
        sample = ""
        for rr in range(start_row, min(start_row + 15, n_rows(grid) + 1)):
            s = cell_str(grid, rr, c)
            if s:
                sample = s[:14]; break
        label = f"{idx_to_col(c)} : {sample}" if sample else f"{idx_to_col(c)} :"
        opts.append(label); o2i[label] = c
    return opts, o2i, {v: k for k, v in o2i.items()}


# ---------- UI ----------
st.set_page_config(page_title="GIS 주소 변환기", page_icon="🌐", layout="centered")

st.markdown("""
<style>
/* 전체 배경 살짝 하늘톤 그라데이션 */
.stApp { background: linear-gradient(180deg, #f3f8fc 0%, #e8f1f9 100%); }
.block-container { padding-top: 2.2rem; max-width: 820px; }

/* 제목 */
h1 { color: #2c3e50; font-weight: 800; }

/* 기능 선택 카드 라디오 */
div[role="radiogroup"] { gap: 12px; flex-wrap: nowrap; }
div[role="radiogroup"] > label {
  flex: 1 1 0; min-width: 0; text-align: center; background: #ffffff;
  border: 2px solid #d6e4f0; border-radius: 16px; padding: 16px 10px;
  transition: all .15s ease; box-shadow: 0 1px 3px rgba(80,120,160,.08); cursor: pointer;
}
div[role="radiogroup"] > label p { white-space: nowrap; }
div[role="radiogroup"] > label:hover { border-color: #9cc4e4; background: #f6fbff; }
div[role="radiogroup"] > label:has(input:checked) {
  background: #cce2f3; border-color: #6fa8d4;
  box-shadow: 0 4px 12px rgba(111,168,212,.30);
}
/* 라디오 동그라미 숨기기 */
div[role="radiogroup"] > label > div:first-child { display: none; }
div[role="radiogroup"] label p { font-size: 1.02rem; font-weight: 700; color: #2c3e50; }

/* 버튼 */
div.stButton > button {
  border-radius: 12px; font-weight: 700; border: none;
  background: #6fa8d4; color: white; padding: .5rem 1rem;
}
div.stButton > button:hover { background: #5b93c2; color: white; }
div[data-testid="stDownloadButton"] > button {
  border-radius: 12px; font-weight: 700; background: #ffffff;
  border: 2px solid #6fa8d4; color: #2c6a9a;
}

/* 카드 컨테이너(테두리) 부드럽게 */
div[data-testid="stVerticalBlockBorderWrapper"] {
  background: #ffffff; border-radius: 16px;
  border: 1px solid #dbe7f1 !important; box-shadow: 0 2px 10px rgba(80,120,160,.06);
}
</style>
""", unsafe_allow_html=True)

st.title("🌐 GIS 주소 변환기")
st.caption("VWorld 기반 · 주소를 PNU·좌표·지도 레이어로 일괄 변환합니다")

with st.expander("ℹ️  사용 방법 (처음이라면 펼쳐 보세요)", expanded=False):
    st.markdown(
        "1. **VWorld 인증키** — 각자 발급(무료) 후 입력합니다.\n"
        "2. **기능 선택** — ① PNU · ② 좌표 · ③ QGIS 레이어\n"
        "3. **파일 업로드** — .xlsx 또는 .csv 파일을 올리면 주소 열을 자동 인식합니다 (미리보기로 확인·수정 가능).\n"
        "4. **변환 시작** — 결과를 표로 확인하고 내려받습니다.\n\n"
        "※ 입력한 키와 파일은 변환에만 쓰이며, 별도로 저장하지 않습니다.")

# 1) 인증키
with st.container(border=True):
    st.markdown("##### 🔑 VWorld 인증키")
    api_key = st.text_input("인증키", value="", type="password",
                            placeholder="본인이 발급받은 VWorld 인증키를 붙여넣어 주세요",
                            label_visibility="collapsed")
    st.caption("발급: [vworld.kr](https://www.vworld.kr) → 오픈API → 인증키 발급 "
               "(활용 API에 **2D 데이터 API** 체크, 사이트 URL은 `http://localhost`)")

# 2) 기능 선택 (카드)
st.markdown("##### 기능 선택")
func = st.radio("기능", ["①  주소 → PNU", "②  주소 → 좌표", "③  QGIS 레이어"],
                horizontal=True, label_visibility="collapsed")

# 3) 옵션
with st.container(border=True):
    want_jiga = want_pt = want_pg = False
    crs_label = list(CRS_OPTIONS)[0]
    if func.startswith("①"):
        st.markdown("주소를 19자리 PNU로 변환합니다 · 결과에 PNU·정제주소·본번·부번 포함")
        want_jiga = st.checkbox("공시지가(원/㎡)·기준연월도 함께 조회합니다 (조금 느려질 수 있습니다)")
    elif func.startswith("②"):
        st.markdown("주소를 지도 좌표로 변환합니다")
        crs_label = st.selectbox("좌표계", list(CRS_OPTIONS))
    else:
        st.markdown("주소를 QGIS에서 바로 열리는 지도 레이어(GeoJSON)로 만듭니다")
        want_pt = st.checkbox("포인트 레이어 — 주소를 점으로 (빠릅니다)", value=True)
        want_pg = st.checkbox("필지 경계 레이어 — 실제 땅 모양·공시지가 포함 (느릴 수 있습니다)", value=True)

# 4) 파일 업로드 + 열 확인
uploaded = st.file_uploader(
    "엑셀(.xlsx) 또는 CSV 파일을 올려 주세요  ·  주소가 담긴 파일", type=["xlsx", "csv"])

if uploaded:
    ext = uploaded.name.rsplit(".", 1)[-1].lower()
    sheet_name = None
    size_mb = (uploaded.size or 0) / 1_000_000
    limit = CSV_SIZE_LIMIT_MB if ext == "csv" else XLSX_SIZE_LIMIT_MB
    if size_mb > limit:
        st.error(
            f"파일이 너무 큽니다 (약 {size_mb:.0f}MB · 현재 한도 {limit}MB).\n\n"
            "이 도구는 **주소 목록**용이에요. 수십만 행짜리 큰 파일은 무료 서버 메모리 한도를 "
            "넘어 멈출 수 있어, 처리를 막았습니다.\n\n"
            "👉 주소가 담긴 부분만 남겨 가볍게 만든 뒤 다시 올려 주세요."
            + ("\n\n💡 엑셀(.xlsx)보다 **CSV로 저장**하면 더 큰 파일도 올릴 수 있습니다." if ext == "xlsx" else ""))
        st.stop()

    truncated = False
    if ext == "csv":
        enc_label = st.selectbox(
            "CSV 인코딩", list(CSV_ENCODINGS),
            help="한글이 깨져 보이면 인코딩을 바꿔 주세요. 공공데이터·SGIS CSV는 보통 CP949입니다.")
        try:
            grid, used_enc, truncated = read_csv_grid(uploaded.getvalue(), CSV_ENCODINGS[enc_label])
        except Exception as e:
            st.error(f"CSV를 읽는 중 문제가 발생했습니다. 인코딩을 바꿔 보세요.\n\n{type(e).__name__}: {e}")
            st.stop()
        if enc_label == "자동 감지":
            st.caption(f"인코딩 자동 감지: **{used_enc}** · 한글이 깨지면 위에서 직접 선택해 주세요.")
    else:
        try:
            sheets = xlsx_sheet_names(uploaded.getvalue())
        except Exception as e:
            st.error(f"파일을 여는 중 문제가 발생했습니다.\n\n{type(e).__name__}: {e}")
            st.stop()
        sheet_name = sheets[0] if sheets else None
        if len(sheets) > 1:
            sheet_name = st.selectbox(
                f"📑 시트 선택 (총 {len(sheets)}개)", sheets,
                help="엑셀에 시트가 여러 개입니다. 변환할 시트를 하나 고르세요. "
                     "시트별로 골라 각각 변환·내려받기 하면 됩니다.")
        try:
            grid, truncated = read_xlsx_grid(uploaded.getvalue(), sheet_name)
        except Exception as e:
            st.error(f"시트를 여는 중 문제가 발생했습니다.\n\n{type(e).__name__}: {e}")
            st.stop()

    if truncated:
        st.warning(f"행이 매우 많아 처음 {MAX_ROWS:,}행까지만 읽었습니다. "
                   "나머지 행은 파일을 나눠 다시 올려 주세요.")

    det = detect_layout(grid)
    sig = f"{sheet_name or 'csv'}-{n_cols(grid)}"   # 시트/구조 바뀌면 열 선택 위젯을 새로 시작

    with st.container(border=True):
        st.markdown("##### 📋 주소 열 확인")
        st.caption("먼저 아래 **파일 미리보기**에서 주소·본번·부번이 각각 어느 열(A·B·C…)에 있는지 확인한 뒤, 그 아래에서 열을 지정하세요.")
        pv_r = min(10, n_rows(grid)); pv_c = min(n_cols(grid), MAX_COLS)
        head_df = pd.DataFrame(
            [[cell_str(grid, rr, c) for c in range(1, pv_c + 1)] for rr in range(1, pv_r + 1)],
            columns=[idx_to_col(c) for c in range(1, pv_c + 1)])
        head_df.index = range(1, pv_r + 1)   # 실제 행 번호(1부터) — '데이터 시작 행' 판단에 도움
        st.dataframe(head_df, use_container_width=True, height=min(38 * (pv_r + 1), 400))
        start_row = st.number_input("데이터 시작 행", min_value=1, value=int(det["start_row"]))
        opts, o2i, i2o = build_options(grid, start_row)
        mode = st.radio("주소 형태", ["한 칸에 전체주소", "여러 칸으로 쪼갬"],
                        index=0 if det["mode"] == "single" else 1, horizontal=True)

        if mode == "한 칸에 전체주소":
            default = i2o.get(det["cols"][0], "(없음)") if det["mode"] == "single" else "(없음)"
            addr_label = st.selectbox("주소 열", opts, index=opts.index(default))
            sel = {"start_row": int(start_row), "kind": "full", "addr_col": o2i[addr_label]}
        else:
            # '한 칸 전체주소'로 감지된 경우엔 그 열을 구성1 기본값으로 넣어 준다(하이브리드 대응)
            admin_default = det["cols"][:-1] if det["mode"] == "split" else det["cols"]
            jibun_default = det["cols"][-1] if det["mode"] == "split" else None
            st.markdown("**주소 구성 열** — 큰 단위 → 작은 단위 순서 (예: 시도·시군구·읍면동·리)")
            st.caption("💡 주소가 한 칸(예: D열 '전북특별자치도 전주시 덕진구 우아동3가')에 통째로 있으면 "
                       "**구성 1에 그 열 하나만** 고르고, 본번·부번은 아래 '본번·부번 분리'로 따로 지정하면 됩니다.")
            cc = st.columns(3); admin_cols = []
            for i in range(5):
                dflt = i2o.get(admin_default[i], "(없음)") if i < len(admin_default) else "(없음)"
                lab = cc[i % 3].selectbox(f"구성 {i+1}", opts, index=opts.index(dflt), key=f"adm{i}_{sig}")
                if o2i[lab]:
                    admin_cols.append(o2i[lab])
            jkind = st.radio("지번 형태", ["한 칸 (71-2)", "본번·부번 분리"], horizontal=True)
            sel = {"start_row": int(start_row), "kind": "split", "admin_cols": admin_cols,
                   "jibun_kind": "cell" if jkind.startswith("한") else "bonbu"}
            if sel["jibun_kind"] == "cell":
                dflt = i2o.get(jibun_default, "(없음)")
                sel["jibun_col"] = o2i[st.selectbox("지번 열", opts, index=opts.index(dflt))]
            else:
                c1, c2 = st.columns(2)
                sel["bon_col"] = o2i[c1.selectbox("본번 열", opts)]
                sel["bu_col"] = o2i[c2.selectbox("부번 열", opts)]

        prefix_common = st.text_input(
            "주소 앞에 공통으로 붙일 내용 (선택)", value="",
            placeholder="예: 전북특별자치도 — 데이터에 시도가 빠져 있을 때만",
            help="파일 전체가 같은 지역인데 주소에 그 앞부분(보통 시도)이 빠져 있으면 여기에 적어 주세요. "
                 "모든 행 앞에 자동으로 붙습니다. 지역이 섞인 파일이면 비워 두세요.").strip()

        prev, r = [], int(start_row)
        while r <= n_rows(grid) and len(prev) < 10:
            a = full_addr(grid, r, sel, prefix_common)
            if a and any(ch.isdigit() for ch in a):
                prev.append(a)
            r += 1
        st.markdown(
            "<div style='background:#efe2df;border-radius:12px;padding:12px 16px;color:#5a4a44'>"
            "<b>미리보기</b><br>" + ("<br>".join(f"· {p}" for p in prev) if prev else "주소 열을 선택해 주세요")
            + "</div>", unsafe_allow_html=True)

    run = st.button("🚀 변환 시작", type="primary", use_container_width=True)

    if run:
        api_key = api_key.strip()   # 붙여넣기 때 딸려온 공백·줄바꿈 제거 (인증키오류 방지)
        if not api_key:
            st.error("먼저 VWorld 인증키를 입력해 주세요.")
            st.stop()

        # (1) 주소 문자열을 먼저 모두 만든다 (로컬 계산, 빠름)
        rows = range(int(start_row), n_rows(grid) + 1)
        addrs = [full_addr(grid, rr, sel, prefix_common) for rr in rows]
        valid = [a for a in addrs if a and any(ch.isdigit() for ch in a)]
        skip = len(addrs) - len(valid)
        crs = CRS_OPTIONS[crs_label][0] if func.startswith("②") else "EPSG:4326"

        if not valid:
            st.warning("변환할 주소를 찾지 못했습니다. 주소 열과 시작 행을 확인해 주세요.")
            st.stop()

        # (2) 한 주소가 필요로 하는 모든 조회(지오코딩+필지)를 한 작업으로 묶는다
        def work(addr):
            pnu, x, y, refined, status = geocode(addr, api_key, crs)
            item = {"addr": addr, "status": status, "pnu": pnu, "x": x, "y": y, "refined": refined}
            need_parcel = ((func.startswith("①") and want_jiga) or
                           (func.startswith("③") and want_pg))
            if need_parcel and status == "OK" and x:
                item["geom"], item["props"] = get_parcel(x, y, api_key)
            return item

        # (3) 여러 주소를 동시에 처리 (병렬) — 결과는 입력 순서대로 보존
        results = [None] * len(valid)
        bar = st.progress(0.0, text="변환 중...")
        total = len(valid)
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            fut2pos = {ex.submit(work, a): p for p, a in enumerate(valid)}
            done = 0
            for fut in as_completed(fut2pos):
                results[fut2pos[fut]] = fut.result()
                done += 1
                bar.progress(done / total, text=f"변환 중... {done}/{total}")
        bar.progress(1.0, text="완료")

        # (4) 순서대로 결과 조립
        records, pts, pgs = [], [], []
        ok = fail = 0
        for item in results:
            addr, status = item["addr"], item["status"]
            pnu, x, y, refined = item["pnu"], item["x"], item["y"], item["refined"]
            if status == "OK":
                ok += 1
            else:
                fail += 1

            if func.startswith("①"):
                st_ = status
                if st_ == "OK" and not (pnu and len(pnu) == 19):
                    st_ = "PNU불완전"
                bon, bu = parse_bonbu(pnu)
                rec = {"입력주소": addr, "PNU": pnu, "정제주소": refined,
                       "본번": bon, "부번": bu, "상태": st_}
                if want_jiga and status == "OK":
                    props = item.get("props") or {}
                    jg = props.get("jiga")
                    rec["공시지가(원/㎡)"] = int(jg) if jg and str(jg).isdigit() else None
                    rec["기준연월"] = (f"{props.get('gosi_year','')}.{props.get('gosi_month','')}".strip(".")
                                    if props else None)
                records.append(rec)
            elif func.startswith("②"):
                xl, yl = CRS_OPTIONS[crs_label][1], CRS_OPTIONS[crs_label][2]
                records.append({"입력주소": addr, xl: x, yl: y, "정제주소": refined, "상태": status})
            else:
                if status == "OK" and x:
                    bon, bu = parse_bonbu(pnu)
                    records.append({"입력주소": addr, "PNU": pnu, "lon": float(x), "lat": float(y), "상태": status})
                    if want_pt:
                        pts.append({"type": "Feature",
                                    "geometry": {"type": "Point", "coordinates": [float(x), float(y)]},
                                    "properties": {"입력주소": addr, "정제주소": refined, "PNU": pnu, "본번": bon, "부번": bu}})
                    if want_pg:
                        geom, props = item.get("geom"), item.get("props") or {}
                        if geom:
                            jg = props.get("jiga")
                            pgs.append({"type": "Feature", "geometry": geom,
                                        "properties": {"정제주소": refined, "PNU": props.get("pnu") or pnu,
                                                       "공시지가": int(jg) if jg and str(jg).isdigit() else None}})
                else:
                    records.append({"입력주소": addr, "상태": status})

        st.session_state["res"] = {"func": func, "records": records, "pts": pts, "pgs": pgs,
                                   "ok": ok, "fail": fail, "skip": skip}
        st.session_state.setdefault("history", []).insert(0, {
            "시각": datetime.now().strftime("%H:%M:%S"), "기능": func[:1],
            "파일": uploaded.name + (f" · {sheet_name}" if sheet_name else ""),
            "성공": ok, "실패": fail, "건너뜀": skip})

if "res" in st.session_state:
    res = st.session_state["res"]
    ok_n, fail_n = res["ok"], res["fail"]
    if ok_n == 0 and fail_n > 0:
        st.error(f"❌ 변환에 모두 실패했습니다 (실패 {fail_n}건). 아래 표의 **상태** 열을 확인하세요.\n\n"
                 "· **통신실패:연결실패/빈응답/502** → VWorld에 연결이 막힌 것 "
                 "(해외 클라우드 IP 차단). 한국 IP 호스팅(Cloudtype)의 주소로 접속했는지 확인하세요.\n"
                 "· **인증키오류** → 키 값·공백·상태 확인.")
    elif fail_n > 0:
        st.warning(f"⚠️ 변환은 끝났지만 일부 실패했습니다 (성공 {ok_n} / 실패 {fail_n}). 상태 열을 확인하세요.")
    else:
        st.success(f"✅ 변환이 완료되었습니다. (성공 {ok_n}건)")
    m1, m2, m3 = st.columns(3)
    m1.metric("✅ 성공", res["ok"]); m2.metric("⚠️ 실패", res["fail"]); m3.metric("➖ 건너뜀", res["skip"])
    df = pd.DataFrame(res["records"])
    st.dataframe(df, use_container_width=True, height=300)
    if {"lat", "lon"}.issubset(df.columns):
        st.map(df.dropna(subset=["lat", "lon"])[["lat", "lon"]])

    func = res["func"]
    if func.startswith("③"):
        if res["pts"]:
            st.download_button("⬇ 포인트 레이어 (.geojson)",
                               json.dumps({"type": "FeatureCollection", "features": res["pts"]}, ensure_ascii=False),
                               file_name="포인트.geojson", mime="application/geo+json", use_container_width=True)
        if res["pgs"]:
            st.download_button("⬇ 필지 경계 레이어 (.geojson)",
                               json.dumps({"type": "FeatureCollection", "features": res["pgs"]}, ensure_ascii=False),
                               file_name="필지.geojson", mime="application/geo+json", use_container_width=True)
        st.caption("내려받은 .geojson 파일을 QGIS 창에 끌어다 놓으면 바로 표시됩니다.")
    else:
        st.markdown("##### ⬇ 결과 내려받기")
        dl1, dl2 = st.columns(2)

        # 1) 엑셀(.xlsx) — 한글 걱정 없음
        buf = io.BytesIO(); df.to_excel(buf, index=False)
        dl1.download_button("엑셀 (.xlsx)", buf.getvalue(), file_name="변환결과.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)

        # 2) CSV(.csv) — 인코딩 선택 (한글 안 깨지게)
        CSV_OUT = {
            "UTF-8 (엑셀·메모장 어디서나 안 깨짐, 권장)": "utf-8-sig",
            "CP949 / EUC-KR (한글 윈도우 엑셀 전용)": "cp949",
        }
        enc_label = dl2.selectbox("CSV 인코딩", list(CSV_OUT),
                                  help="엑셀에서 열었을 때 한글이 깨지면 다른 인코딩으로 받아 보세요. "
                                       "보통 UTF-8이면 됩니다.")
        csv_bytes = df.to_csv(index=False).encode(CSV_OUT[enc_label], errors="replace")
        dl2.download_button("CSV (.csv)", csv_bytes, file_name="변환결과.csv",
                            mime="text/csv", use_container_width=True)

# ---- 세션 변환 기록 ----
if st.session_state.get("history"):
    st.markdown("##### 📜 변환 기록")
    st.dataframe(pd.DataFrame(st.session_state["history"]),
                 use_container_width=True, hide_index=True)
    st.caption("※ 이 기록은 현재 브라우저 세션에만 임시 저장됩니다 (새로고침하면 사라집니다).")
